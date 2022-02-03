import openpyxl
import os
import psycopg2
from datetime import datetime, timedelta
from dicts import endpoints_cells

PG_HOST, PG_PORT, PG_DATABASE, PG_USER, PG_PASS = os.environ['PG01_CRED'].split(':')
conn = psycopg2.connect(dbname=PG_DATABASE, user=PG_USER, password=PG_PASS, host=PG_HOST, port=PG_PORT)
cur = conn.cursor()

wb = openpyxl.load_workbook('ПЛАН на Январь 22г.xlsx')
dwb = openpyxl.load_workbook('ПЛАН на Январь 22г.xlsx', data_only=True)

# now_date = datetime.strptime('2022-01-19 00:00:00', '%Y-%m-%d %H:%M:%S')
start_date = input("Введите дату (Формат %Y-%m-%d %H:%M:%S) (Enter, если сегодня): ")
try:
    start_date = datetime.strptime(start_date, '%Y-%m-%d %H:%M:%S')
except:
    start_date = datetime.now()
next_date = start_date + timedelta(days=1)

for key in endpoints_cells.keys():
    place = key.split(" ")
    if place[1] == "ЦПБФ":
        d_wb = dwb['ЦПБФ']
        f_wb = wb['ЦПБФ']
    else:
        d_wb = dwb['ЦПКФ']
        f_wb = wb['ЦПКФ']

    r = int(place[0].replace('A', '')) - 1
    v = endpoints_cells.get(key)
    if v[2].strip() == str(d_wb[f'A{r}'].value).strip():
        dates = d_wb[f'W{r}:AO{r}'][0]
        for date in dates:
            if str(date.value) == str(start_date):
                for i in range(1, 5):
                    cell_v = d_wb.cell(row=date.row + i, column=date.column)
                    if cell_v.value:
                        if int(cell_v.value) < 0:
                            plan = 0
                        else:
                            plan = cell_v.value

                        cell_f = f_wb.cell(row=date.row + i, column=date.column)
                        if str(cell_f.value).find('G') != -1:
                            plan_info = '{"ТО":' + str(round(d_wb[f'E{date.row + i}'].value)) + '}'
                        else:
                            plan_info = None

                        br = (str(d_wb[f'B{date.row + i}'].value).strip()).capitalize()
                        cur.execute(f""" SELECT shift_start, shift_end
                                         FROM source.shift_calendar
                                         WHERE shift_start >= '{start_date}' AND shift_start < '{next_date}'
                                         AND brigada = '{br}' AND schedule_id = (SELECT schedule_id
                                                                                 FROM parameters.endpoint_schedules
                                                                                 WHERE endpoint_id = {v[0]});""")
                        resp = cur.fetchall()

                        print(79, v[0], resp[0][0], resp[0][1], round(cell_v.value), plan_info)
                        # cur.execute(""" INSERT INTO source.production_plan (client_id, endpoint_id, period_start,
                        #                                                     period_end, plan, plan_info)
                        #                 VALUES (%s, %s, %s, %s, %s, %s);""", (79, v[0], resp[0][0], resp[0][1], round(plan),
                        #                                                       plan_info))
                        # conn.commit()

cur.close()
conn.close()
